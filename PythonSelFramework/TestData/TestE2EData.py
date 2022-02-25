import openpyxl


class TestE2EData:

    @staticmethod
    def getTestData(test_case_name):
        dictionary = dict()
        book = openpyxl.load_workbook('C:\\DEV\\Python\\AUTOMATION\\Framework\\PythonSelFramework\\PythonDemo.xlsx')
        sheet = book.active
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == test_case_name:
                for column in range(2, sheet.max_column + 1):
                    dictionary[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
        return [dictionary]
